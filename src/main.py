import os
import sys
import glob
import argparse
import pandas as pd
import numpy as np
from openpyxl.styles import Font

# ベスト10実績Excelに含まれるシステム用シート名（店舗データではないシート）
SYSTEM_SHEET_NAMES = frozenset([
    "今期部門売上", "今期売上実績", "コードなど",
    "前期ベスト10に対して今期順位", "今期ベスト10に対して前期順位",
    "前期部門売上", "前期売上実績　データ",
])


def find_latest_file(pattern):
    """
    指定されたパターンにマッチするファイルの中から、最終更新日時が最も新しいファイルを返す。
    """
    files = glob.glob(pattern)
    if not files:
        return None
    return max(files, key=os.path.getmtime)


def get_valid_store_sheets(xls_best):
    """
    ExcelFileオブジェクトからシステム用シートを除いた店舗シート名リストを返す。
    シートの並び順はExcelファイル内の順序を保持する。
    """
    return [s for s in xls_best.sheet_names if s not in SYSTEM_SHEET_NAMES]


def process_store(best10_file, target_store, current_sales_map, output_dir):
    """
    指定された1店舗分のベスト10データをパースし、今月売上と突合した結果Excelを出力する。

    Args:
        best10_file: ベスト10実績Excelファイルパス
        target_store: 処理対象の店舗シート名
        current_sales_map: 商品コードをキーとする販売管理表の辞書
        output_dir: 出力先ディレクトリ
    """
    print(f"\nベスト10基準データ（{target_store}シート）を読み込み中...")
    df_base = pd.read_excel(best10_file, sheet_name=target_store)

    dept_data = {}
    dept_order = []
    current_dept = None

    for _idx, row in df_base.iterrows():
        val0 = row.iloc[0]
        nan_count = row.isna().sum()

        # 部門名の判定
        if pd.notna(val0) and nan_count >= len(row) - 3:
            val0_str = str(val0).strip()
            if val0_str not in ["順位", "今期", "合計", "ベスト10小計"] and not val0_str.isdigit():
                current_dept = val0_str
                if current_dept not in dept_data:
                    dept_data[current_dept] = []
                    dept_order.append(current_dept)
                continue

        if current_dept is None:
            continue

        try:
            rank = int(val0)
            if 1 <= rank <= 10:
                r = row.tolist()

                code_str = ""
                if pd.notna(r[2]):
                    try:
                        if isinstance(r[2], float) and r[2].is_integer():
                            code_str = str(int(r[2]))
                        else:
                            code_str = str(r[2]).strip().split('.')[0]
                    except (ValueError, AttributeError):
                        code_str = str(r[2]).strip()

                # 今月の実績値を取得
                current_sales = np.nan
                current_ratio = np.nan
                if code_str in current_sales_map:
                    current_sales = current_sales_map[code_str]["sales"]
                    current_ratio = current_sales_map[code_str]["ratio"]

                dept_data[current_dept].append({
                    "今期順位": rank,
                    "前期順位": pd.to_numeric(r[1], errors='coerce'),
                    "商品コード": code_str,
                    "商品名": str(r[3]).strip() if pd.notna(r[3]) else "",
                    "前年売上実績": pd.to_numeric(r[4], errors='coerce'),
                    "前年比較日比(%)": pd.to_numeric(r[8], errors='coerce'),
                    "今月売上実績": current_sales,
                    "今月比較日比(%)": current_ratio,
                    "前年打数実績": pd.to_numeric(r[9], errors='coerce'),
                    "前年一品単価": pd.to_numeric(r[13], errors='coerce'),
                })
        except (ValueError, TypeError):
            pass

    # ExcelWriter を使用して複数シートのExcelファイルを作成
    output_excel = os.path.join(output_dir, f"best10_with_current_sales_{target_store}.xlsx")
    print(f"結果Excelファイルを作成中: {output_excel}...")

    red_bold_font = Font(color="FF0000", bold=True)

    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        for dept in dept_order:
            items = dept_data[dept]
            if not items:
                continue
            df_dept = pd.DataFrame(items)

            # シート名の文字数制限への対応
            sheet_name = dept
            for char in [':', '\\', '/', '?', '*', '[', ']']:
                sheet_name = sheet_name.replace(char, '_')
            sheet_name = sheet_name[:31]

            df_dept.to_excel(writer, sheet_name=sheet_name, index=False)

            worksheet = writer.sheets[sheet_name]

            # %カラムのインデックスを特定（1-indexed for openpyxl）
            percent_columns = []
            header_row = [cell.value for cell in worksheet[1]]
            for col_idx, header in enumerate(header_row, start=1):
                if header is not None and "(%)" in str(header):
                    percent_columns.append(col_idx)

            # %カラムに小数点第1位の数値フォーマットを適用
            for row_idx in range(2, worksheet.max_row + 1):
                for col_idx in percent_columns:
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.0'

            # 売上金額・単価カラムに整数フォーマット（小数点なし）を適用
            integer_format_columns = []
            integer_target_headers = {"前年売上実績", "今月売上実績", "前年一品単価"}
            for col_idx, header in enumerate(header_row, start=1):
                if header is not None and str(header) in integer_target_headers:
                    integer_format_columns.append(col_idx)

            for row_idx in range(2, worksheet.max_row + 1):
                for col_idx in integer_format_columns:
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0'

            # 赤太字条件付き書式の適用
            # 「今月比較日比(%)」と「商品名」のカラム位置をヘッダーから動的に特定
            current_ratio_col = None
            product_name_col = None
            for col_idx, header in enumerate(header_row, start=1):
                if header == "今月比較日比(%)":
                    current_ratio_col = col_idx
                elif header == "商品名":
                    product_name_col = col_idx

            if current_ratio_col is not None and product_name_col is not None:
                for row_idx in range(2, worksheet.max_row + 1):
                    cell_ratio = worksheet.cell(row=row_idx, column=current_ratio_col)
                    ratio_val = cell_ratio.value

                    if ratio_val is not None and isinstance(ratio_val, (int, float)) and ratio_val <= 100:
                        cell_name = worksheet.cell(row=row_idx, column=product_name_col)
                        cell_name.font = red_bold_font
                        cell_ratio.font = red_bold_font

            # 列幅の自動調整
            for col in worksheet.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value is not None:
                        val_str = str(cell.value)
                        str_len = sum(2 if ord(c) > 0x7F else 1 for c in val_str)
                        max_len = max(max_len, str_len)
                worksheet.column_dimensions[col_letter].width = max(max_len + 3, 10)

    print(f"突合Excelレポートの作成（店舗: {target_store}）が完了しました。")


def load_sales_data(sales_file):
    """
    販売管理表を読み込み、商品コードをキーとする売上・比較日比の辞書を返す。

    Returns:
        dict: {商品コード(str): {"sales": float, "ratio": float}}
    """
    print("最新販売管理表を読み込み中...")
    df_sales_raw = pd.read_excel(sales_file, sheet_name="Sheet1", skiprows=6, header=None)

    current_sales_map = {}
    for _idx, row in df_sales_raw.iterrows():
        code = row.iloc[0]

        if pd.isna(code) or str(code).strip() in ["合計", "総計"] or "対象期間" in str(code):
            continue

        try:
            if isinstance(code, float) and code.is_integer():
                code_str = str(int(code))
            else:
                code_str = str(code).strip().split('.')[0]
        except (ValueError, AttributeError):
            code_str = str(code).strip()

        if not code_str.isdigit():
            continue

        sales_amt = pd.to_numeric(row.iloc[3], errors='coerce')
        ratio_val = pd.to_numeric(row.iloc[6], errors='coerce')

        current_sales_map[code_str] = {
            "sales": sales_amt if pd.notna(sales_amt) else 0,
            "ratio": ratio_val if pd.notna(ratio_val) else np.nan,
        }

    print(f"販売管理表から {len(current_sales_map)} 件の商品をマッピングしました。")
    return current_sales_map


def main():
    # 日本語出力の文字化けを防ぐ
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')

    # コマンドライン引数の解析
    parser = argparse.ArgumentParser(description="ベスト10実績と今月売上の突合ツール")
    parser.add_argument(
        "--store",
        "-s",
        default="全店",
        help="参照する基準Excelの店舗シート名（デフォルト: 全店）。'all' を指定すると全店舗を一括出力",
    )
    args = parser.parse_args()
    target_store = args.store.strip()

    input_dir = "input_data"
    output_dir = "output_data"

    # フォルダの自動作成
    os.makedirs(input_dir, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)

    # input_data フォルダから最新のExcelファイルを自動検出
    best10_pattern = os.path.join(input_dir, "ベスト１０実績_*.xlsx")
    sales_pattern = os.path.join(input_dir, "販売管理表(単品別売上実績)-*.xlsx")

    best10_file = find_latest_file(best10_pattern)
    sales_file = find_latest_file(sales_pattern)

    if not best10_file:
        print(f"Error: '{input_dir}' フォルダ内に 'ベスト１０実績_*.xlsx' が見つかりません。")
        sys.exit(1)
    if not sales_file:
        print(f"Error: '{input_dir}' フォルダ内に '販売管理表(単品別売上実績)-*.xlsx' が見つかりません。")
        sys.exit(1)

    print(f"使用するベスト10基準ファイル: {best10_file}")
    print(f"使用する最新販売管理表ファイル: {sales_file}")

    # Excelファイルを開いて店舗シートを検証
    try:
        xls_best = pd.ExcelFile(best10_file)
        valid_stores = get_valid_store_sheets(xls_best)
    except Exception as exc:
        print(f"Excelファイルの検証中にエラーが発生しました: {exc}")
        sys.exit(1)

    # 処理対象の店舗リストを決定
    if target_store == "all":
        stores_to_process = valid_stores
        print(f"\n全シート一括出力モード: {len(stores_to_process)} 店舗を処理します")
        print(f"対象店舗: {', '.join(stores_to_process)}")
    else:
        if target_store not in xls_best.sheet_names:
            print(f"\n[ERROR] 指定された店舗シート '{target_store}' はExcelファイル内に存在しません。")
            print(f"利用可能な店舗名: {', '.join(valid_stores)}")
            sys.exit(1)
        stores_to_process = [target_store]
        print(f"参照店舗（シート名）: {target_store}")

    # 販売管理表の読み込み（全店舗で共通のため1回のみ）
    current_sales_map = load_sales_data(sales_file)

    # 各店舗を順次処理
    success_count = 0
    error_stores = []
    for store_name in stores_to_process:
        try:
            process_store(best10_file, store_name, current_sales_map, output_dir)
            success_count += 1
        except Exception as exc:
            print(f"\n[WARNING] 店舗 '{store_name}' の処理中にエラーが発生しました: {exc}")
            error_stores.append(store_name)

    # 処理結果サマリーの表示
    print("\n" + "=" * 60)
    if target_store == "all":
        print(f"全シート一括出力 完了: {success_count}/{len(stores_to_process)} 店舗を正常に出力しました。")
        if error_stores:
            print(f"[WARNING] エラーが発生した店舗: {', '.join(error_stores)}")
    else:
        print(f"突合Excelレポートの作成（店舗: {target_store}）が完了しました。")
    print(f"出力先: {output_dir}/")
    print("=" * 60)


if __name__ == "__main__":
    main()
